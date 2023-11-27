import pandas as pd
from openpyxl import load_workbook
import xlrd
import numpy as np
import os
import sys
from file_preprocess import *


def get_merged_cells_view(data_path, sheet_name, dataframe):
    file_extension = os.path.splitext(data_path)[1]
    if file_extension in ['.xlsm', '.xlsx']:
        return get_merged_cells_view_xlsx(data_path, sheet_name, dataframe)
    if file_extension == '.xls':
        return get_merged_cells_view_xls(data_path, sheet_name, dataframe)


def get_merged_cells_view_xlsx(data_path, sheet_name, dataframe):
    workbook = load_workbook(data_path)
    sheet = workbook[sheet_name]  # Используем переданное название страницы
    merged_cells_ranges = sheet.merged_cells.ranges
    for merged_cell_range in merged_cells_ranges:
        merged_cell_value = sheet.cell(row=merged_cell_range.min_row, column=merged_cell_range.min_col).value
        rows_slice = slice(merged_cell_range.min_row - 1, merged_cell_range.max_row - 1)
        cols_slice = slice(merged_cell_range.min_col - 1, merged_cell_range.max_col - 1)
        dataframe.loc[rows_slice, cols_slice] = merged_cell_value
    return dataframe


def get_merged_cells_view_xls(data_path, sheet_name, dataframe):
    workbook = xlrd.open_workbook(data_path, formatting_info=True, ragged_rows=True)

    sheet = workbook.sheet_by_name(sheet_name)
    merged_cells = sheet.merged_cells
    for merged_cell in merged_cells:
        start_row, end_row, start_col, end_col = merged_cell
        cell_value = sheet.cell_value(start_row, start_col)
        rows_slice = slice(start_row, end_row-1)
        cols_slice = slice(start_col, end_col-1)
        dataframe.loc[rows_slice, cols_slice] = cell_value
    return dataframe


def row_is_empty(dataframe_row):
    for value in dataframe_row.values:
        if value != '' and not pd.isnull(value):
            return False
    return True


def get_slice_indexes(dataframe):
    slice_indexes = [dataframe.index[0]]
    prev_empty_row = 0
    for row in dataframe.index[1:]:
        check = row_is_empty(dataframe.loc[row])
        if check:
            if row == (prev_empty_row + 1):
                prev_empty_row = row
                continue
            else:
                slice_indexes.append(row)
                prev_empty_row = row
    slice_indexes.append(dataframe.index[-1])
    return slice_indexes


def get_index_empty_columns(dataframe, threshold=5):
    empty_columns_count = 0
    empty_start_column = None

    for column in dataframe.columns:
        if dataframe[column].isnull().all():
            if empty_columns_count == 0:
                empty_start_column = column
            empty_columns_count += 1
            if empty_columns_count >= threshold:
                return empty_start_column
        else:
            empty_columns_count = 0
            empty_start_column = None

    return None


def remove_spaces(cell_value):
    if str(cell_value).strip() == 'nan':
        return None
    return str(cell_value).strip()


def form_data_slice(dataframe, slice_indexes, header_indexes):
    result = {}
    for index, (start, end) in enumerate(zip(slice_indexes[:-1], slice_indexes[1:]), start=1):
        try:
            matching_num = next(num for num in header_indexes if start <= num <= end)
            result[index] = dataframe.iloc[matching_num:end, :-1].to_json(orient='split', force_ascii=True)
        except StopIteration:
            pass
    return result


def main(path, sheetName):
    dataframe = pd.read_excel(path, sheet_name=sheetName, header=None,
                              keep_default_na=True).applymap(convert_datetime)
    for_processing = dataframe.copy().applymap(remove_spaces)
    check_empty_cols = get_index_empty_columns(dataframe)
    if check_empty_cols is not None:
        for_processing = for_processing.iloc[:, :check_empty_cols]
        dataframe = dataframe.iloc[:, :check_empty_cols]
    merged_cells_view = get_merged_cells_view(path, sheetName, dataframe)
    slice_indexes = get_slice_indexes(merged_cells_view)
    row_index_header, similar_indexes = get_header_indexes(merged_cells_view)
    data_slice = form_data_slice(dataframe, slice_indexes, similar_indexes)
    header_strings = arrayFilter(merged_cells_view.iloc[row_index_header, :-1].values.tolist())
    result = {
        'table': data_slice,
        'header': header_strings,
    }
    json_result = json.dumps(result)
    return json_result


if __name__ == "__main__":
    # main(path=r'C:\Users\vveli\OneDrive\Рабочий стол\1. МЭР, Тех.режимы\Тех.режимы\!4_2014_Тех режим_нефть.xlsm',
    #      sheetName='Нефтяные добывающие')
    file_path = sys.argv[1]
    sheetName = sys.argv[2]
    data = main(file_path, sheetName)
    sys.stdout.write(data)
    sys.stdout.flush()
