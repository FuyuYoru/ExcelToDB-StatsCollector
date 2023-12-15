import json
import re
from openpyxl import load_workbook
import pandas as pd
from datetime import datetime


def get_count_unique_classes(row):
    uniqueStrings = set()
    for idx, cellValue in enumerate(row):
        if isinstance(cellValue, str):
            cleanedValue = re.sub(r'[^a-zA-Zа-яА-Я]', '', cellValue)
            uniqueStrings.add(cleanedValue)
    return list(uniqueStrings)


def get_original_view(data_path, sheet_name):
    dataframe = pd.read_excel(data_path, sheet_name=sheet_name, header=None)
    workbook = load_workbook(data_path)
    sheet = workbook[sheet_name]  # Используем переданное название страницы
    merged_cells_ranges = sheet.merged_cells.ranges
    for merged_cell_range in merged_cells_ranges:
        merged_cell_value = sheet.cell(row=merged_cell_range.min_row, column=merged_cell_range.min_col).value
        rows_slice = slice(merged_cell_range.min_row - 1, merged_cell_range.max_row - 1)
        cols_slice = slice(merged_cell_range.min_col - 1, merged_cell_range.max_col - 1)
        dataframe.loc[rows_slice, cols_slice] = merged_cell_value
    return dataframe


def get_similar_rows(df, index, tolerance_factor=0.5):
    """
    Функция возвращает список индексов строк, близких по количеству уникальных элементов,
    но не одинаковых с данной строкой.
    :param tolerance_factor: float, максимальное отклонение в количестве уникальных элементов в д. ед.
    :param df: DataFrame, в котором производится поиск.
    :param index: int, индекс строки, для которой ищем близкие строки.
    :return: list, список индексов близких строк.
    """

    target_unique_count = len(df['unique_class_values'][index])
    tolerance = int(target_unique_count * tolerance_factor)
    similar_indexes = []
    for idx, unique_values in df['unique_class_values'].items():
        if idx != index and abs(len(unique_values) - target_unique_count) <= tolerance:
            similar_indexes.append(idx)

    return similar_indexes


def get_index_empty_columns(dataframe, threshold=5):
    empty_columns_count = 0
    empty_start_column = None

    for column in dataframe.columns:
        if dataframe[column].isnull().all():
            if empty_columns_count == 0:
                empty_start_column = column
            empty_columns_count += 1
            if empty_columns_count >= threshold:
                return empty_start_column
        else:
            empty_columns_count = 0
            empty_start_column = None

    return None


def get_header_indexes(dataframe):
    dataframe['unique_class_values'] = dataframe.apply(get_count_unique_classes, axis=1)
    row_index_header = dataframe['unique_class_values'].apply(len).idxmax()
    similar_indexes = get_similar_rows(dataframe, row_index_header)
    similar_indexes.append(row_index_header)
    similar_indexes.sort()
    similar_data = dataframe.loc[similar_indexes, dataframe.columns != 'unique_class_values'].values.tolist()
    filtered_indexes = filter_nearest_values(similar_indexes, similar_data)
    return row_index_header, filtered_indexes


def preprocess_and_compare(a, b):
    a_processed = str(a).lower().replace(" ", "")
    b_processed = str(b).lower().replace(" ", "")
    return a_processed == b_processed


def get_blocks(dataframe):
    blocks_json = {}
    block_index = dataframe[dataframe[0] == "<block>"].index
    for num in range(len(block_index) - 1):
        blocks_json[f"block {num + 1}"] = [int(block_index[num]) + 1, int(block_index[num + 1]) - 1]
    blocks_json = json.dumps(blocks_json)
    return blocks_json


def arrayFilter(array):
    temp_array = {}
    for number, item in enumerate(array):
        if isinstance(item, str):
            if item not in temp_array.values():
                temp_array[number + 1] = item
            else:
                temp_array[number + 1] = item + f' col {number}'
        else:
            temp_array[number + 1] = f'undefined_{number + 1}'
    return temp_array


def arrayIsEqual(array1, array2, threshold=0.9):
    if len(array1) != len(array2):
        return False
    count = 0
    for elem1, elem2 in zip(array1, array2):
        if str(elem1).strip() == str(elem2).strip():
            count += 1
    result = (count / len(array1)) >= threshold
    return result


def filter_nearest_values(arrayIndexes, arrayValues):
    if not arrayIndexes:
        return []
    result = [arrayIndexes[0]]
    for i in range(len(arrayIndexes) - 1):
        index_check = abs(arrayIndexes[i] - arrayIndexes[i + 1]) == 1
        values_check = arrayIsEqual(arrayValues[i], arrayValues[i + 1])
        if values_check and index_check:
            result[-1] = arrayIndexes[i + 1]
        if not index_check:
            result.append(arrayIndexes[i + 1])
    return result


def decode_unicode_escapes(data):
    if isinstance(data, str):
        return data.encode('utf-8').decode('unicode-escape')
    elif isinstance(data, list):
        return [decode_unicode_escapes(item) for item in data]
    elif isinstance(data, dict):
        return {key: decode_unicode_escapes(value) for key, value in data.items()}
    else:
        return data


def convert_datetime(value):
    if isinstance(value, (str, datetime)):
        try:
            dt = pd.to_datetime(value)
            date = dt.strftime('%Y-%m')
            return date
        except ValueError:
            return value
    else:
        return value
